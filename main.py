import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile

import urllib.request
import openpyxl
from bs4 import BeautifulSoup


# Read data from Excel file
exceldata = pd.read_excel('input.xlsx', sheet_name='Sheet1')

# Write to Excel
wb = openpyxl.load_workbook("output.xlsx") #workbook Object
ws = wb['Sheet1']  #worksheet object


for idx in exceldata.index:
   # urllib.request.urlopen(exceldata['URLs'][idx]) it will hit the URLs
   # we need to pass that to beautiful Soup to get the response in HTML
   # so urls variable store all the response after hitting URLs from Excel file
   urls = BeautifulSoup(urllib.request.urlopen(exceldata['URLs'][idx]))
   expectedTitle = exceldata['ExpectedOutput'][idx]

   if(expectedTitle == urls.title.string):
      ws['C' + str(idx+2)].value = 'PASS'
      wb.save("output.xlsx")
   else:
      ws['C' + str(idx+2)].value = 'FAIL'
      wb.save("output.xlsx")